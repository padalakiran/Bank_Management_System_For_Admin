import pandas as pd
import xlrd
import openpyxl

def with_dr():

    path='Enter Path Of Excel File'
    df = pd.read_excel(path)
    wb= openpyxl.load_workbook(path)
    ws=wb['Sheet1']
    a=input('Enter Account Number: ')
    b=input('Enter Phone Number: ')



    location = (path)
    wb1= xlrd.open_workbook(location)
    s1 = wb1.sheet_by_index(0)


    s1.cell_value(0,0)                       
    try:
        for i in range(0,s1.nrows+1):
        # print(i)
            if(df.loc[i,'AccountNumber']==int(a)):
                if(df.loc[i,'PhoneNo']==int(b)):
                # print('D{0}'.format(i))
                    c=int(input("Enter amount to Withdraw: "))
                    
                    if (ws['D{0}'.format(i+2)].value-c)<0:
                        print("Insufficiant Balance")
                        print()
                    else:
                        ws['D{0}'.format(i+2)].value=ws['D{0}'.format(i+2)].value-c
                        print('WithDraw Sucussful, Please Collect your Cash....')
                        print()
                    wb.save("data\data.xlsx")
                    
            else:
                if i==s1.nrows:
                      print("user Not Found")
                      break
    
    except:
        pass
#with_dr()
#github: - @padalakiran
